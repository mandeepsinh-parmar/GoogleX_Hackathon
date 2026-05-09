from openpyxl.worksheet import pagebreak
import io
import json
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def generate_pdf_report(park_data: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = styles['Title']
    title_style.textColor = colors.HexColor("#1e3a8a")
    heading_style = styles['Heading2']
    heading_style.textColor = colors.HexColor("#2563eb")
    normal_style = styles['Normal']
    normal_style.fontSize = 10
    normal_style.leading = 14
    bold_style = ParagraphStyle('Bold', parent=normal_style, fontName='Helvetica-Bold')

    story = []

    # Title
    park_name = park_data.get("name", "Industrial Park Report")
    story.append(Paragraph(f"Research Report: {park_name}", title_style))
    story.append(Spacer(1, 20))

    # Basic Info
    state = park_data.get("state", "")
    district = park_data.get("district", "")
    sector = park_data.get("sector", "")
    area = park_data.get("available_area_acres", 0)
    
    story.append(Paragraph("<b>1. General Information</b>", heading_style))
    story.append(Paragraph(f"<b>Location:</b> {district}, {state}", normal_style))
    story.append(Paragraph(f"<b>Primary Sector:</b> {sector.capitalize()}", normal_style))
    story.append(Paragraph(f"<b>Available Land:</b> {area} Acres", normal_style))
    
    # Google Maps Link
    lat = park_data.get("lat")
    lng = park_data.get("lng")
    if lat and lng:
        maps_link = f'<a href="https://www.google.com/maps/search/?api=1&query={lat},{lng}" color="blue">View on Google Maps</a>'
        story.append(Paragraph(f"<b>Map:</b> {maps_link} ({lat}, {lng})", normal_style))
    story.append(Spacer(1, 15))

    # Logistics & Infrastructure
    story.append(Paragraph("<b>2. Logistics & Infrastructure</b>", heading_style))
    hw = park_data.get("nearest_highway_km", "N/A")
    rw = park_data.get("nearest_railway_km", "N/A")
    air = park_data.get("nearest_airport_km", "N/A")
    port = park_data.get("nearest_port_km", "N/A")
    
    data = [
        ["Nearest Highway", f"{hw} km" if hw != "N/A" else hw],
        ["Nearest Railway", f"{rw} km" if rw != "N/A" else rw],
        ["Nearest Airport", f"{air} km" if air != "N/A" else air],
        ["Nearest Port", f"{port} km" if port != "N/A" else port],
        ["Water Supply", park_data.get("water_availability", "Unknown")],
        ["Power Supply", park_data.get("power_availability", "Unknown")],
    ]
    
    t = Table(data, colWidths=[150, 350])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 15))

    # ── Section 3: Comprehensive AI Report ─────────────────────────────────────
    story.append(Paragraph("<b>3. Comprehensive AI Report</b>", heading_style))

    import os
    import requests

    api_key  = os.getenv("GOOGLE_API_KEY", "")
    raw_mats = park_data.get("raw_materials_nearby", [])
    research = park_data.get("research", {})
    schemes  = park_data.get("schemes", {})
    incentives = park_data.get("incentives", [])
    why_suitable   = research.get("why_suitable", "")
    why_attractive = research.get("why_attractive", "")
    spec_benefits  = research.get("specific_benefits", [])

    prompt = (
        f"You are a senior industrial real estate consultant writing a formal investor-grade site report. "
        f"Write exactly 3 paragraphs of plain prose (no bullet points, no markdown asterisks, no headers). "
        f"Use the following data:\n\n"
        f"Park Name: {park_name}\n"
        f"Location: {district}, {state}\n"
        f"Sector: {sector}\n"
        f"Available Land: {area} Acres\n"
        f"Nearest Highway: {hw} km | Nearest Airport: {air} km | "
        f"Nearest Railway: {park_data.get('nearest_railway_km', 'N/A')} km | "
        f"Nearest Port: {park_data.get('nearest_port_km', 'N/A')} km\n"
        f"Water Supply: {park_data.get('water_availability', 'N/A')}\n"
        f"Power Supply: {park_data.get('power_availability', 'N/A')}\n"
        f"Raw Materials Nearby: {', '.join(raw_mats) if raw_mats else 'N/A'}\n"
        f"Park Incentives: {', '.join(incentives) if incentives else 'N/A'}\n"
        f"Estimated Subsidy Stack: Rs. {schemes.get('total_subsidy_cr', 0)} Crore\n"
        f"Estimated Net Investment: Rs. {schemes.get('net_investment_cr', 0)} Crore\n"
        f"AI Research Notes: {why_suitable} {why_attractive}\n"
        f"Specific Benefits: {', '.join(spec_benefits) if spec_benefits else 'N/A'}\n\n"
        f"Paragraph 1 — Strategic Location & Connectivity: Describe why the location is strategically advantageous.\n"
        f"Paragraph 2 — Sector Fit & Raw Material Access: Explain why this park suits the {sector} sector.\n"
        f"Paragraph 3 — Financial Case & Investment Outlook: Summarise the subsidy stack and net investment appeal."
    )

    # v1beta supports all current Gemini models; try in order of preference
    _MODELS = ["gemini-2.0-flash", "gemini-1.5-flash", "gemini-1.0-pro"]
    ai_text    = None
    last_error = None

    for model_name in _MODELS:
        url = (
            f"https://generativelanguage.googleapis.com/v1beta/models/"
            f"{model_name}:generateContent?key={api_key}"
        )
        try:
            payload = {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.4, "maxOutputTokens": 700}
            }
            resp      = requests.post(url, json=payload,
                                      headers={"Content-Type": "application/json"},
                                      timeout=20)
            resp_data = resp.json()
            if resp.status_code == 200 and "candidates" in resp_data:
                ai_text = resp_data["candidates"][0]["content"]["parts"][0]["text"].strip()
                break
            else:
                last_error = resp_data.get("error", {}).get("message", "Unknown error")
        except Exception as exc:
            last_error = str(exc)

    if ai_text:
        # Render each paragraph
        for para in ai_text.split("\n"):
            if para.strip():
                story.append(Paragraph(para.strip(), normal_style))
                story.append(Spacer(1, 7))
    else:
        # ── Rich data-driven narrative (always produces readable content) ───────
        p1 = (
            f"{park_name} is strategically located in {district}, {state}, offering strong "
            f"multimodal connectivity for industrial operations. "
            f"The park sits {hw} km from the nearest highway and {air} km from the nearest airport, "
            f"ensuring efficient logistics for both raw material inflow and finished goods distribution. "
            f"Railway access at {park_data.get('nearest_railway_km', 'N/A')} km and port proximity at "
            f"{park_data.get('nearest_port_km', 'N/A')} km further enhance its export-readiness. "
            f"Reliable water supply via {park_data.get('water_availability', 'local sources')} and "
            f"{park_data.get('power_availability', 'grid power')} make it operationally resilient."
        )
        p2 = (
            f"The park is well-suited for the {sector} sector, with a total of {area} acres of "
            f"available industrial land providing ample room for large-scale manufacturing units. "
        )
        if raw_mats:
            p2 += (
                f"Key raw materials including {', '.join(raw_mats)} are accessible within the region, "
                f"reducing inbound logistics costs and supply-chain risk for {sector} businesses. "
            )
        if incentives:
            p2 += (
                f"The park also offers specific operational advantages such as {', '.join(incentives[:3])}, "
                f"which directly benefit new industrial entrants in this category."
            )

        p3 = (
            f"From a financial standpoint, {park_name} presents a compelling investment case. "
            f"With an estimated subsidy stack of Rs. {schemes.get('total_subsidy_cr', 0)} Crore from "
            f"combined central and state government schemes, the effective net investment is brought down "
            f"to approximately Rs. {schemes.get('net_investment_cr', 0)} Crore. "
            f"This significant subsidy coverage, combined with the park's infrastructure readiness, "
            f"makes it a high-priority site for investors targeting {state} for {sector} expansion."
        )

        for p in [p1, p2, p3]:
            story.append(Paragraph(p, normal_style))
            story.append(Spacer(1, 8))

    story.append(Spacer(1, 15))


    # Schemes & Subsidies
    story.append(Paragraph("<b>4. Subsidies & Government Schemes</b>", heading_style))
    schemes = park_data.get("schemes", {})
    
    story.append(Paragraph(f"<b>Estimated Net Investment:</b> ₹{schemes.get('net_investment_cr', '0')} Crore", bold_style))
    story.append(Paragraph(f"<b>Total Estimated Subsidy:</b> ₹{schemes.get('total_subsidy_cr', '0')} Crore", bold_style))
    story.append(Spacer(1, 10))
    
    scheme_data = [["Type", "Scheme Name", "Estimated Value"]]
    for sch in schemes.get("central_schemes", []):
        scheme_data.append(["Central", Paragraph(sch.get("name", ""), normal_style), f"₹{sch.get('estimated_value_cr', 0)} Cr"])
    for sch in schemes.get("state_schemes", []):
        scheme_data.append(["State", Paragraph(sch.get("name", ""), normal_style), f"₹{sch.get('estimated_value_cr', 0)} Cr"])
        
    if len(scheme_data) > 1:
        st = Table(scheme_data, colWidths=[60, 340, 100])
        st.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(st)
    else:
        story.append(Paragraph("No specific schemes found.", normal_style))

    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_report(park_data: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Park Report"
    
    # Headers
    headers = [
        "Park Name", "State", "District", "Sector", "Available Land (Acres)", 
        "Latitude", "Longitude", "Map Link", "Nearest Highway (km)", 
        "Nearest Railway (km)", "Nearest Airport (km)", "Nearest Port (km)", 
        "Water Availability", "Power Availability", "Raw Materials", 
        "Net Investment (Cr)", "Total Subsidy (Cr)"
    ]
    
    ws.append(headers)
    
    # Format Headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Data Row
    lat = park_data.get("lat", "")
    lng = park_data.get("lng", "")
    map_link = f"https://www.google.com/maps/search/?api=1&query={lat},{lng}" if lat and lng else ""
    schemes = park_data.get("schemes", {})
    
    row = [
        park_data.get("name", ""),
        park_data.get("state", ""),
        park_data.get("district", ""),
        park_data.get("sector", ""),
        park_data.get("available_area_acres", ""),
        lat,
        lng,
        map_link,
        park_data.get("nearest_highway_km", ""),
        park_data.get("nearest_railway_km", ""),
        park_data.get("nearest_airport_km", ""),
        park_data.get("nearest_port_km", ""),
        park_data.get("water_availability", ""),
        park_data.get("power_availability", ""),
        ", ".join(park_data.get("raw_materials_nearby", [])),
        schemes.get("net_investment_cr", 0),
        schemes.get("total_subsidy_cr", 0)
    ]
    
    ws.append(row)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
